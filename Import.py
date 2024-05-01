from Objects import *

nr_of_players = 0


def import_all(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)

    import_players(wb['Players'])
    import_games(wb['Games'])
    import_preferences(wb['Prefs'])
    # Import the Berger tabel to initiate a Round Robin pairing.
    import_berger(wb["Berger" + str(len(Player.players))])


def import_players(ws_players):
    for i in range(2, ws_players.max_row + 1):
        Player(ws_players.cell(row=i, column=1).value, ws_players.cell(row=i, column=2).value)
    global nr_of_players
    nr_of_players = len(Player.players)


def import_games(ws_games):
    for i in range(2, ws_games.max_row + 1):
        Game(ws_games.cell(row=i, column=1).value)


# Imports the Prefs table
def import_preferences(ws_preferences):
    for i in range(2, ws_preferences.max_column + 1):
        for j in range(2, ws_preferences.max_row + 1):
            Preference(ws_preferences.cell(row=j, column=i).value,
                       get_player_by_name(ws_preferences.cell(row=1, column=i).value),
                       get_game_by_name(ws_preferences.cell(row=j, column=1).value))


def import_berger(ws_berger):
    for i in range(1, nr_of_players):
        round_i = Round(i)
        for j in range(1, int(nr_of_players / 2) + 1):
            pairing_nrs = ws_berger.cell(row=i, column=j).value.split(':')
            round_i.matches.append(Match(int(pairing_nrs[0]), int(pairing_nrs[1])))
