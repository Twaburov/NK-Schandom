from Objects import *


def export_all(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    wb.remove(wb['Pairing'])
    print_pairing(wb.create_sheet('Pairing'))
    wb.save(filename)


def print_pairing(ws):
    rownr = 1

    for r in Round.rounds:
        matchnr = 0
        for m in r.matches:
            ws.cell(row=rownr, column=1, value=str(r.round_nr) + chr(65 + matchnr))  # e.g: 1A, 3F, 5D
            ws.cell(row=rownr, column=2, value=Player.player_dict[m.p1nr].name)
            ws.cell(row=rownr, column=3, value=Player.player_dict[m.p2nr].name)
            if m.Game != None:
                ws.cell(row=rownr, column=6,
                        value=m.Game.name)  # The name of the game and then some information about how big the weight was for each player.
                ws.cell(row=rownr, column=7,
                        value=sum(p.weight for p in get_player_by_pairing_number(m.p1nr).playerprefs if p.Game == m.Game))
                ws.cell(row=rownr, column=8,
                        value=sum(p.weight for p in get_player_by_pairing_number(m.p2nr).playerprefs if p.Game == m.Game))
            rownr += 1
            matchnr += 1
